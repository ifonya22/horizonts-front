import aiomysql
import asyncio
from utils.shared_data import (
    USER_,
    PASSWORD_,
    HOST_,
    PORT_,
    DB_,
    SLEEP_,
    ALPHABET_,
)
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import os


# Функция обработки фильтров
async def options(
    objects, date_s, time_s, date_e, time_e, krit=False, prost=False
):
    if len(objects) == 0:
        return  # Отсутствуют объекты
    if date_s is None:
        date_s = datetime.now().strftime("%Y-%m-%d")
    if time_s is None:
        time_s = "00:00:01"
    if date_e is None:
        date_e = datetime.now().strftime("%Y-%m-%d")
    if time_e is None:
        time_e = "23:59:59"
    fetch_data(objects, date_s, time_s, date_e, time_e, krit, prost)
    return


#  Обработка типа процесса
def type_processing(type_stc):
    if type_stc == "Krit":
        return "Критический"
    elif type_stc == "Prost":
        return "Простой"
    else:
        return "Нормальный"


# Сбор данныз
async def fetch_data(objects, date_s, time_s, date_e, time_e, krit, prost):
    pool = await aiomysql.create_pool(
        user=USER_, password=PASSWORD_, host=HOST_, port=PORT_, db=DB_
    )
    result = []  # Список для хранения результатов запросов
    async with pool.acquire() as connection:
        async with connection.cursor(aiomysql.DictCursor) as cursor:
            placeholders = ", ".join(["%s"] * len(objects))
            if prost and krit:
                query = f"""
                    SELECT date_stc, time_stc, power_stc, id_obj_stc, type_stc
                    FROM statistics
                    WHERE id_obj_stc IN ({placeholders})
                    AND date_stc BETWEEN %s AND %s
                    AND time_stc BETWEEN %s AND %s
                    AND (type_stc='Krit' OR type_stc='Prost')
                    ORDER BY id_obj_stc, date_stc, time_stc
                """
                params = objects + [date_s, date_e, time_s, time_e]
            elif prost:
                query = f"""
                    SELECT date_stc, time_stc, power_stc, id_obj_stc, type_stc
                    FROM statistics
                    WHERE id_obj_stc IN ({placeholders})
                    AND date_stc BETWEEN %s AND %s
                    AND time_stc BETWEEN %s AND %s
                    AND type_stc='Prost'
                    ORDER BY id_obj_stc, date_stc, time_stc
                """
                params = objects + [date_s, date_e, time_s, time_e]
            elif krit:
                query = f"""
                    SELECT date_stc, time_stc, power_stc, id_obj_stc, type_stc
                    FROM statistics
                    WHERE id_obj_stc IN ({placeholders})
                    AND date_stc BETWEEN %s AND %s
                    AND time_stc BETWEEN %s AND %s
                    AND type_stc='Krit'
                    ORDER BY id_obj_stc, date_stc, time_stc
                """
                params = objects + [date_s, date_e, time_s, time_e]
            else:
                query = f"""
                    SELECT date_stc, time_stc, power_stc, id_obj_stc, type_stc
                    FROM statistics
                    WHERE id_obj_stc IN ({placeholders})
                    AND date_stc BETWEEN %s AND %s
                    AND time_stc BETWEEN %s AND %s
                    ORDER BY id_obj_stc, date_stc, time_stc
                """
                params = objects + [date_s, date_e, time_s, time_e]

            # Выполнение запроса с передачей параметров
            await cursor.execute(query, params)
            objects = await cursor.fetchall()
            # Добавление результатов запроса в общий список
            for el in objects:
                result.append(
                    {
                        "time_stc": datetime.strptime(
                            str(el["time_stc"]), "%H:%M:%S"
                        ),
                        "date_stc": el["date_stc"],
                        "power_stc": el["power_stc"],
                        "id_obj_stc": el["id_obj_stc"],
                        "type_stc": type_processing(el["type_stc"]),
                    }
                )

    # Закрытие пула соединений
    pool.close()
    await pool.wait_closed()
    return result


# Данные по оборудованиям
async def get_data_object(objects):
    pool = await aiomysql.create_pool(
        user=USER_, password=PASSWORD_, host=HOST_, port=PORT_, db=DB_
    )
    async with pool.acquire() as connection:
        async with connection.cursor(aiomysql.DictCursor) as cursor:
            placeholders = ", ".join(["%s"] * len(objects))
            query = f"""
                        SELECT * FROM objects
                        WHERE id_obj IN ({placeholders})
                    """
            # Выполнение запроса с передачей параметров
            await cursor.execute(query, objects)
            return await cursor.fetchall()


# Обработка данных для диаграмм
def generate_rows(data, objects):
    datetime_power_map = {}

    for row in data:
        # Формируем ключ как комбинацию даты и времени
        datetime_str = (
            row["date_stc"],
            row["time_stc"].strftime("%H:%M:%S"),
        )  # Преобразуем время в строку

        if datetime_str not in datetime_power_map:
            datetime_power_map[datetime_str] = {}

        datetime_power_map[datetime_str][row["id_obj_stc"]] = row["power_stc"]

    # Формируем заголовок таблицы
    rows = [["Date", "Time"] + [f"Object {i+1}" for i in range(len(objects))]]
    last_date = ""
    # Заполняем таблицу данными
    for (date, time), values in sorted(datetime_power_map.items()):
        if last_date != date:
            row = [
                date.strftime("%Y-%m-%d"),
                time,
            ]  # Преобразуем дату в строку
            # Добавляем дату и время
            last_date = date
        else:
            row = ["", time]

        for obj in objects:
            row.append(
                values.get(obj, rows[-1][2])
            )  # Если данных нет, подставляем 0
        rows.append(row)
    return rows


# Создание отчётов
async def generate_excel_report(
    objects, date_s, time_s, date_e, time_e, krit=False, prost=False
):
    # Получаем данные с помощью fetch_data и get_data_object
    data = await fetch_data(
        objects, date_s, time_s, date_e, time_e, krit, prost
    )
    data_obj = await get_data_object(objects)

    # Если данные есть, начинаем формирование отчета
    if data and data_obj:
        os.makedirs("reports", exist_ok=True)
        output_file = f"reports/report {date_s}-{date_e}.xlsx"
        # Создаем новую книгу и активируем лист
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Отчет"

        # Устанавливаем заголовоки
        title = "Отчет по электроэнергии"
        ws.merge_cells("A1:E1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        title = f"Для цеха №{data_obj[0]['workshop_id']}. За отчётный период c {date_s} по {date_e}"
        ws.merge_cells("A2:E2")
        ws["A2"] = title
        ws["A2"].font = Font(size=12)
        ws["A2"].alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )

        # Заголовки колонок
        ws["A3"] = "Дата"
        ws["B3"] = "Время"
        ws["C3"] = "Мощность"
        ws["D3"] = "Тип"
        ws["E3"] = "Оборудование"
        # Делаем заголовки жирными
        for col in ["A", "B", "C", "D", "E"]:
            ws[f"{col}3"].font = Font(bold=True)

        # Заполняем таблицу данными
        row = 4
        for record in data:
            obj = next(
                (o for o in data_obj if o["id_obj"] == record["id_obj_stc"]),
                None,
            )
            ws[f"A{row}"] = record["date_stc"]
            ws[f"B{row}"] = record["time_stc"].strftime("%H:%M:%S")
            ws[f"C{row}"] = record["power_stc"]
            ws[f"D{row}"] = record["type_stc"]
            ws[f"E{row}"] = record["id_obj_stc"]
            row += 1

        # Ширина столбца по max(cell)
        max_widths = [0] * 5
        for row_idx in range(3, 5):
            for col_idx in range(5):
                cell_value = ws[
                    get_column_letter(col_idx + 1) + str(row_idx)
                ].value
                if cell_value:
                    max_widths[col_idx] = max(
                        max_widths[col_idx], len(str(cell_value))
                    )

        # Применяем ширину колонок
        for col_idx, width in enumerate(max_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 5

        # Устанавливаем выравнивание для всех ячеек
        for row in range(3, ws.max_row + 1):
            for col in ["A", "B", "C", "D", "E"]:  # с 4 и до конца
                cell = ws[f"{col}{row}"]
                cell.alignment = Alignment(horizontal="right")
        for col in ["A", "B", "C", "D", "E"]:  # 3 строка
            cell = ws[f"{col}3"]
            cell.alignment = Alignment(horizontal="center")

        # Сохранение данных об оборудовании
        ws = wb.create_sheet("Оборудование")
        # Устанавливаем заголовоки
        title = "Исходные данные оборудования"
        ws.merge_cells("A1:D1")
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A2"] = "Номер оборудования"
        ws["B2"] = "Критическая граница"
        ws["C2"] = "Граница холостого хода"
        ws["D2"] = "Холостой ход"

        # Делаем заголовки жирными
        for col in ["A", "B", "C", "D"]:
            ws[f"{col}2"].font = Font(bold=True)

        # Заполняем
        row = 3
        for record in data_obj:

            ws[f"A{row}"] = record["id_obj"]
            ws[f"B{row}"] = record["max_obj"]
            ws[f"C{row}"] = record["max_obj"] * record["percent_obj"]
            ws[f"D{row}"] = record["percent_obj"]
            row += 1

        # Ширина столбца по max(cell)
        max_widths = [0] * 4
        for row_idx in range(2, 3):
            for col_idx in range(4):
                cell_value = ws[
                    get_column_letter(col_idx + 1) + str(row_idx)
                ].value
                if cell_value:
                    max_widths[col_idx] = max(
                        max_widths[col_idx], len(str(cell_value))
                    )

        # Применяем ширину колонок
        for col_idx, width in enumerate(max_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

        # Устанавливаем выравнивание для всех ячеек
        for row in range(3, ws.max_row + 1):
            for col in ["A", "B", "C", "D", "E"]:  # с 3 и до конца
                cell = ws[f"{col}{row}"]
                cell.alignment = Alignment(horizontal="right")
        for col in ["A", "B", "C", "D", "E"]:  # 2 строка
            cell = ws[f"{col}2"]
            cell.alignment = Alignment(horizontal="center")

        # Статистика
        ws = wb.create_sheet("Статистика")

        rows = generate_rows(data, objects)
        if not rows or not objects:
            print("Нет данных для графика.")
            return

        # Заполняем
        row = 3
        for record in rows:
            for i in range(0, len(record)):
                ws[f"{ALPHABET_[i]}{row}"] = record[i]
            row += 1
        # Создаем линейный график
        chart = LineChart()

        # Устанавливаем заголовоки
        title = "Статистика электрооборудования"
        q = "A1:" + ALPHABET_[len(objects) + 1] + "1"
        ws.merge_cells(q)
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")
        title = f"Для цеха №{data_obj[0]['workshop_id']}. За отчётный период c {date_s} по {date_e}"

        q = "A2:" + ALPHABET_[len(objects) + 1] + "2"
        ws.merge_cells(q)
        ws["A2"] = title
        ws["A2"].font = Font(size=12)
        ws["A2"].alignment = Alignment(
            horizontal="center", vertical="center", wrapText=True
        )

        ws["A3"] = "Дата"
        ws["B3"] = "Время"
        for i in range(len(objects)):
            ws[f"{ALPHABET_[i+2]}3"] = f"Оборудование {objects[i]}"
        # Ширина столбца по max(cell)
        max_widths = [0] * (objects[-1] + 2)
        for row_idx in range(3, 5):
            for col_idx in range(5):
                cell_value = ws[
                    get_column_letter(col_idx + 1) + str(row_idx)
                ].value
                if cell_value:
                    max_widths[col_idx] = max(
                        max_widths[col_idx], len(str(cell_value))
                    )

        # Применяем ширину колонок
        for col_idx, width in enumerate(max_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width + 2

        # Устанавливаем выравнивание для всех ячеек
        for row in range(3, ws.max_row + 1):
            for col in ALPHABET_[0 : len(objects) + 1]:
                cell = ws[f"{col}{row}"]
                cell.alignment = Alignment(horizontal="right")
        for col in ALPHABET_[0 : len(objects) + 1]:  # 3 строка
            cell = ws[f"{col}3"]
            cell.alignment = Alignment(horizontal="center")
        chart.title = "Линейная диаграмма"
        chart.y_axis.title = "Мощность, кВт/ч"
        chart.x_axis.title = "Дата"

        # Определяем диапазоны данных для графика
        data_range = Reference(
            ws,
            min_col=3,
            max_col=3 + len(objects) - 1,
            min_row=3,
            max_row=row - 1,
        )
        categories_range = Reference(
            ws, min_col=1, max_col=1, min_row=4, max_row=row - 1
        )
        chart.add_data(data_range, titles_from_data=True)
        chart.set_categories(categories_range)
        ws.add_chart(chart, f"{ALPHABET_[len(objects)+4]}2")
        # Сохраняем файл
        wb.save(output_file)

        print(f"Отчет сохранен в {output_file}")
        return output_file
    else:
        print("Нет данных для отчета.")

    return None


# Пример использования
if __name__ == "__main__":
    date_s = "2024-09-03"
    time_s = "00:00:00"
    date_e = "2024-09-03"
    time_e = "23:59:59"
    objects = [2, 3]

    # Запускаем асинхронную функцию generate_excel_report с заданными параметрами
    asyncio.run(generate_excel_report(objects, date_s, time_s, date_e, time_e))
